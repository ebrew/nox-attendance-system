from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, logout as logout_check, login as login_checks
from django.contrib import messages
from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
# from .tokens import account_activation_token
from django.utils.http import urlsafe_base64_decode
from django.contrib.auth import get_user_model
# from maintenance_portal.settings import EMAIL_HOST_USER
from django.core.mail import send_mail
from django.template.loader import render_to_string
# from printer_support.emails import send_pending_feedback_email
from django.contrib.auth.signals import user_logged_in, user_logged_out
import socket
from datetime import datetime, timedelta
from django.db.models import Avg
from django.shortcuts import render
from .forms import *
from django.contrib.auth import get_user_model
from .models import *
import datetime as dt
import openpyxl

User = get_user_model()

internet_issues = (OSError, socket.gaierror)


def is_connected():
    try:
        socket.create_connection(("1.1.1.1", 53))
        return True
    except internet_issues:
        return False


# Person.objects.using('icps_db').create(...)
# Person.objects.using('icps_db').create(...)
# user.save(using='icps_db')
# check = User.objects.all().using('icps_db')
#     print('Test started: ', check)


def home(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            # ICPSAccess.objects.all().delete()
            # ICPSAccessUniqueAttendance.objects.all().delete()
            # ICPSEmployee.objects.all().delete()
            return render(request, 'users/home.html')
        return render(request, "users/home.html")
    else:
        return render(request, 'users/login.html')


def about(request):
    return render(request, 'about.html')


@login_required(login_url='login')
def logout(request):
    u = request.user
    logout_check(request)
    messages.success(request, 'Logged out successfully {}! '
                              'Thanks for spending some quality time with the Web site today.'.format(u))
    return render(request, 'users/login.html')


def login(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid:
            email = request.POST['email']
            password = request.POST['password']

            try:
                existing_user = User.objects.get(email=email)
                inactive_user = User.objects.filter(email=email, is_active=False)
            except (TypeError, ValueError, OverflowError, User.DoesNotExist):
                existing_user = None
            if existing_user is None:
                print(existing_user)
                messages.warning(request, 'Your account does not exist! Fill the below form to join the team.')
                return render(request, 'users/register.html')
            elif inactive_user:
                messages.warning(request, 'Your account is inactive! Wait for admin approval to login')
                return render(request, 'users/login.html')
            elif existing_user:
                user = authenticate(email=email, password=password)
                if user is not None:
                    if user.is_active and user.is_staff:
                        login_checks(request, user)
                        # Event.objects.create(user_id=user.pk, action='Logged in as ADMIN')
                        return redirect('home')
                    elif user.is_active:
                        login_checks(request, user)
                        # Event.objects.create(user_id=user.pk, action='Logged in as TECHNICIAN')
                        return redirect('home')
                else:
                    messages.warning(request, 'Invalid credentials')
                    return render(request, 'users/login.html')
    else:
        form = LoginForm()
    return render(request, 'users/login.html', {'form': form})


def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid:
            access = request.POST['access']
            p1 = request.POST['password1']
            p2 = request.POST['password2']
            email = request.POST['email']
            fn = request.POST['first_name']
            ln = request.POST['last_name']
            inactive_user = User.objects.filter(email=email, is_active=False)

            try:
                valid_password = (p1 == p2)
                existing_user = User.objects.get(email=email, is_active=True)
            except (TypeError, ValueError, OverflowError, User.DoesNotExist):
                existing_user = None
            if not valid_password:
                messages.warning(request, 'Password mismatch!')
                return render(request, 'users/register.html')
            elif inactive_user:
                messages.warning(request, 'Account already exist! Wait for admin approval to login')
                return render(request, 'users/login.html')
            elif existing_user:
                messages.warning(request, 'Account already exist!')
                return render(request, 'users/login.html')
            else:
                if access == 'admin':
                    user = User.objects.create_user(email=email, first_name=fn, last_name=ln, password=p1,
                                                    is_staff=True)
                else:
                    user = User.objects.create_user(email=email, first_name=fn, last_name=ln, password=p1)
                user.save()
                messages.success(request, 'Account created successfully!')
                return render(request, 'data_import.html')
    else:
        form = RegisterForm()
    return render(request, 'users/register.html', {'form': form})


# uploading file for storage
def upload_file(request):
    if request.method == 'POST':
        excel_file = request.FILES["excel_file"]
        area = request.POST['area']

        # subsidiary selection validation
        if area == "none":
            messages.warning(request, f'Select the subsidiary to upload to')
            return render(request, 'data_import.html')

        invalid_data = False
        excel_data = list()  # array format

        # validations to check extension or file size
        if str(excel_file)[-5:] != ".xlsx":
            messages.warning(request, f'{excel_file} is NOT AN EXCEL FILE!!!')
            return render(request, 'data_import.html')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheets = wb.sheetnames  # getting all sheets list of strings
            print('Available sheets: ', len(sheets))
            if len(sheets) > 1:
                messages.warning(request, f'{excel_file} has multiple sheets!')
                return render(request, 'data_import.html')

            # getting only the first sheet(sheet 1)
            worksheet = wb["Sheet1"]

            # iterating over the rows and getting value from each cell in row
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))

                excel_data.append(row_data)  # db entries in a list format

            # validation on the first column heading
            if len(excel_data[0][0]) != 19:
                messages.warning(request, f'{excel_file} has invalid data format, No Column heading is required, '
                                          f'only the data!!!')
                return render(request, 'data_import.html')

        except KeyError:
            invalid_data = True

        if invalid_data:
            messages.warning(request, f'{excel_file} has invalid data format, '
                                      f'Consider renaming the first sheet as "Sheet1"')
            return render(request, 'data_import.html')

        if area == "icps":

            # verifying area group
            if str(excel_data[0][0])[2] != ".":
                messages.warning(request, f'Sorry, {excel_file} seem to belong to Head Office Group!!!')
                return render(request, 'data_import.html')

            # ready to save to db
            for entry in excel_data:
                obj = ICPSAccess.objects.create(
                    created_at=dt.datetime.strptime(entry[0], '%d.%m.%Y %H:%M:%S'),
                    date=dt.datetime.strptime(entry[0], '%d.%m.%Y %H:%M:%S').date(),
                    morpho_device=entry[1],
                    name=entry[2],
                    key=entry[3],
                    access=entry[4]
                )
                obj.save()

                # Adds new employees
                employee = ICPSEmployee.objects.filter(name__iexact=entry[2])
                if not employee.exists():
                    obj2 = ICPSEmployee.objects.create(name=entry[2])
                    obj2.save()
            messages.success(request, f'{excel_file} data saved successfully to ICPS group!')
        else:

            # verifying area group
            if str(excel_data[0][0])[2] != "/":
                messages.warning(request, f'Sorry, {excel_file} seem to belong to ICPS Group!!!')
                return render(request, 'data_import.html')

            for entry in excel_data:
                obj = Access.objects.create(
                    created_at=dt.datetime.strptime(entry[0], '%d/%m/%Y %H:%M:%S'),
                    date=dt.datetime.strptime(entry[0], '%d/%m/%Y %H:%M:%S').date(),
                    first_name=entry[1],
                    last_name=entry[2],
                    user_policy=entry[3],
                    morpho_device=entry[5],
                    key=entry[6],
                    access=entry[7]
                )
                if entry[4] != 'None':
                    obj.employee_id = entry[4],
                obj.save()

                # Adds new employees
                employee = Employee.objects.filter(first_name__iexact=entry[1], last_name__iexact=entry[2])
                if not employee.exists():
                    obj2 = Employee.objects.create(first_name=entry[1], last_name=entry[2], department=entry[3])
                    if entry[4] != 'None':
                        obj2.employee_id = entry[4],
                    obj2.save()
            messages.success(request, f'{excel_file} data saved successfully to Head Office group!')
        return render(request, 'data_import.html')

    return render(request, 'data_import.html')


@login_required(login_url='login')
def options(request):
    return render(request, 'report_options.html')


# ICPS employees
@login_required(login_url='login')
def icps_employees(request):
    users = ICPSEmployee.objects.all().order_by('name')
    index = 0
    for i in users:
        index += 1
        i.comment = index
    return render(request, "icps/icps_employees.html", {'users': users})


def convert_to_hour_minute(time_value):
    if time_value is not None:
        # Convert to string and remove fractional seconds part if present
        time_string = str(time_value).split('.')[0]

        # Parse the time string
        time_object = datetime.strptime(time_string, "%H:%M:%S")

        # Format the time as "hour:minute"
        formatted_time = time_object.strftime("%H:%M")

        return formatted_time
    else:
        return None


# Head Office employees
@login_required(login_url='login')
def octagon_employees(request):
    users = Employee.objects.all().order_by('first_name')
    # client = Employee.objects.all().order_by('first_name').using('octagon_db')
    index = 0
    for i in users:
        index += 1
        i.comment = index
    return render(request, "octagon/octagon_employees.html", {'users': users})


# Delete an employee
@login_required(login_url='login')
def delete_icps_employee(request, pk):
    current_user = request.user
    user = ICPSEmployee.objects.get(id=pk)

    if request.method == 'POST':
        ICPSEmployee.objects.get(id=pk).delete()
        messages.success(request, f"Employee {user} deleted successfully!")
        return redirect('icps_employees')
    return render(request, 'icps/icps_delete_prompt.html', {'item': user})


# Delete an employee
@login_required(login_url='login')
def delete_octagon_employee(request, pk):
    current_user = request.user
    user = Employee.objects.get(id=pk)

    if request.method == 'POST':
        Employee.objects.get(id=pk).delete()
        messages.success(request, f"Employee {user} deleted successfully!")
        return redirect('octagon_employees')
    return render(request, 'octagon/octagon_delete_prompt.html', {'item': user})


@login_required(login_url='login')
def employees_report(request, ready):
    title, start_date, end_date, employees, employee_unique_dates, comment = '', '', '', [], [], []
    summary_ready = False

    if ready == 'False':
        messages.warning(request, 'Date Range required!')
        return render(request, "report.html",
                      {'employees': employee_unique_dates, 'title': title, 'start_date': start_date,
                       'end_date': end_date, 'summary_ready': ready})

    elif request.method == 'POST':
        summary_ready = True
        start_date = request.POST['date']
        end_date = request.POST["date2"]
        title = f"From {dt.datetime.strptime(start_date, '%Y-%m-%d').strftime('%d %b, %Y')} to " \
                f"{dt.datetime.strptime(end_date, '%Y-%m-%d').strftime('%d %b, %Y')}"

        access = Access.objects.filter(date__range=(start_date, end_date)) \
            .exclude(first_name="(unknown user)").values_list('date', 'first_name', 'last_name') \
            .distinct().order_by('-date', 'first_name', 'last_name')

        for i in access:
            employee_last_out = Access.objects.filter(date=i[0], first_name=i[1], last_name=i[2]). \
                latest('created_at')  # user time out
            employee_first_in = Access.objects.filter(date=i[0], first_name=i[1], last_name=i[2]). \
                earliest('created_at')  # user time in
            employee_last_out.timein = str(employee_first_in.created_at.time())
            employee_last_out.timeout = str(employee_last_out.created_at.time())
            employee_last_out.employee_id = Employee.objects.get(first_name=i[1], last_name=i[2]).id

            # user time out and time in difference
            t1 = dt.datetime.strptime(employee_last_out.timein, "%H:%M:%S")
            t2 = dt.datetime.strptime(employee_last_out.timeout, "%H:%M:%S")

            # checking and deducting breaktime
            t = (t2 - t1)
            if (t.total_seconds() / 3600) >= 1:
                t = (t2 - t1) - timedelta(hours=1)
            else:
                t = (t2 - t1)
            # checking for overtime

            if (t.total_seconds() / 3600) >= 8:
                employee_last_out.ot = t - timedelta(hours=8)
                employee_last_out.comment = 'Sufficient work time'
            else:
                employee_last_out.ot = dt.timedelta()
                if not employee_last_out.comment:
                    employee_last_out.comment = 'Insufficient work time'

            employee_last_out.total = convert_to_hour_minute(t)  # total user work hours
            employee_last_out.regular = convert_to_hour_minute(t - employee_last_out.ot)
            employee_unique_dates.append(employee_last_out)

    return render(request, "report.html", {'employees': employee_unique_dates, 'title': title, 'start_date': start_date,
                                           'end_date': end_date, 'summary_ready': summary_ready})


# Head office
@login_required(login_url='login')
def employee_report_summary(request, start_date, end_date, title):
    employees = Employee.objects.all().exclude(first_name="(unknown user)").order_by('first_name', 'last_name')
    index = 0
    for user in employees:
        index += 1
        user.employee_id = index
        user.ot = dt.timedelta()
        user.total = dt.timedelta()
        user.regular = dt.timedelta()

        access = Access.objects.filter(date__range=(start_date, end_date), first_name=user.first_name,
                                       last_name=user.last_name).values_list('date', 'first_name', 'last_name') \
            .distinct().order_by('-date')

        for i in access:
            employee_last_out = Access.objects.filter(date=i[0], first_name=i[1], last_name=i[2]). \
                latest('created_at')  # user time out
            employee_first_in = Access.objects.filter(date=i[0], first_name=i[1], last_name=i[2]). \
                earliest('created_at')  # user time in
            employee_last_out.timein = str(employee_first_in.created_at.time())
            employee_last_out.timeout = str(employee_last_out.created_at.time())

            # user time out and time in difference
            t1 = dt.datetime.strptime(employee_last_out.timein, "%H:%M:%S")
            t2 = dt.datetime.strptime(employee_last_out.timeout, "%H:%M:%S")

            # checking and deducting breaktime
            t = (t2 - t1)
            if (t.total_seconds() / 3600) >= 1:
                t = (t2 - t1) - timedelta(hours=1)
            else:
                t = (t2 - t1)

            # checking for overtime
            if (t.total_seconds() / 3600) >= 8:
                employee_last_out.ot = t - timedelta(hours=8)
                user.ot += employee_last_out.ot
            else:
                employee_last_out.ot = dt.timedelta()
                user.ot += employee_last_out.ot

            employee_last_out.total = t  # total user work hours
            employee_last_out.regular = t - employee_last_out.ot
            user.regular += employee_last_out.regular
            user.total += employee_last_out.total

    return render(request, "octagon/summery_report.html", {'employees': employees, 'title': title})


@login_required(login_url='login')
def icps_employees_report(request, ready):
    title, start_date, end_date, employees, employee_unique_dates, comment = '', '', '', [], [], []
    summary_ready = False

    if ready == 'False':
        messages.warning(request, 'Start and End Date required!')
        return render(request, "icps/icps_report.html",
                      {'employees': employee_unique_dates, 'title': title, 'start_date': start_date,
                       'end_date': end_date, 'summary_ready': ready})

    elif request.method == 'POST':
        summary_ready = True
        start_date = request.POST['date']
        end_date = request.POST["date2"]
        title = f"From {dt.datetime.strptime(start_date, '%Y-%m-%d').strftime('%d %b, %Y')} to " \
                f"{dt.datetime.strptime(end_date, '%Y-%m-%d').strftime('%d %b, %Y')}"

        access = ICPSAccess.objects.filter(date__range=(start_date, end_date)).values_list('date', 'name') \
            .distinct().order_by('name', '-date')

        for i in access:
            employee_last_out = ICPSAccess.objects.filter(date=i[0], name=i[1]).latest('created_at')  # user time out
            employee_first_in = ICPSAccess.objects.filter(date=i[0], name=i[1]).earliest('created_at')  # user time in
            employee_last_out.timein = str(employee_first_in.created_at.time())
            employee_last_out.timeout = str(employee_last_out.created_at.time())
            employee_last_out.employee_id = ICPSEmployee.objects.get(name=i[1]).id

            # user time out and time in difference
            t1 = dt.datetime.strptime(employee_last_out.timein, "%H:%M:%S")
            t2 = dt.datetime.strptime(employee_last_out.timeout, "%H:%M:%S")

            # checking and deducting breaktime
            t = (t2 - t1)
            if (t.total_seconds() / 3600) >= 1:
                t = (t2 - t1) - timedelta(hours=1)
            else:
                t = (t2 - t1)
                employee_last_out.comment = 'Insufficient work time, Failed access procedures'

            # checking for overtime
            if (t.total_seconds() / 3600) >= 8:
                employee_last_out.ot = t - timedelta(hours=8)
                employee_last_out.comment = 'Sufficient work time'
            else:
                employee_last_out.ot = dt.timedelta()
                if not employee_last_out.comment:
                    employee_last_out.comment = 'Insufficient work time'

            employee_last_out.total = convert_to_hour_minute(t)  # total user work hours
            employee_last_out.regular = convert_to_hour_minute(t - employee_last_out.ot)
            employee_unique_dates.append(employee_last_out)

    return render(request, "icps/icps_report.html",
                  {'employees': employee_unique_dates, 'title': title, 'start_date': start_date,
                   'end_date': end_date, 'summary_ready': summary_ready})


# View detailed maintenance occurrence
@login_required(login_url='login')
def icps_employee_report_summary(request, start_date, end_date, title):
    employees = ICPSEmployee.objects.all().order_by('name')
    index = 0
    for user in employees:
        index += 1
        user.employee_id = index
        user.ot = dt.timedelta()
        user.total = dt.timedelta()
        user.regular = dt.timedelta()

        access = ICPSAccess.objects.filter(date__range=(start_date, end_date), name=user.name). \
            values_list('date', 'name').distinct().order_by('-date')

        for i in access:
            employee_last_out = ICPSAccess.objects.filter(date=i[0], name=i[1]).latest('created_at')  # user time out
            employee_first_in = ICPSAccess.objects.filter(date=i[0], name=i[1]).earliest('created_at')  # user time in
            employee_last_out.timein = str(employee_first_in.created_at.time())
            employee_last_out.timeout = str(employee_last_out.created_at.time())

            # user time out and time in difference
            t1 = dt.datetime.strptime(employee_last_out.timein, "%H:%M:%S")
            t2 = dt.datetime.strptime(employee_last_out.timeout, "%H:%M:%S")

            # checking and deducting breaktime
            t = (t2 - t1)
            if (t.total_seconds() / 3600) >= 1:
                t = (t2 - t1) - timedelta(hours=1)
            else:
                t = (t2 - t1)

            # checking for overtime
            if (t.total_seconds() / 3600) >= 8:
                employee_last_out.ot = t - timedelta(hours=8)
                user.ot += employee_last_out.ot
            else:
                employee_last_out.ot = dt.timedelta()
                user.ot += employee_last_out.ot

            employee_last_out.total = t  # total user work hours
            employee_last_out.regular = t - employee_last_out.ot
            user.regular += employee_last_out.regular
            user.total += employee_last_out.total

    return render(request, "icps/icps_summery_report.html", {'employees': employees, 'title': title})


@login_required(login_url='login')
def icps_employees_avg_report(request, ready):
    title, start_date, end_date, users, employee_unique_dates, comment = '', '', '', [], [], []
    summary_ready = False

    if ready == 'False':
        messages.warning(request, 'Start and End Date required!')
        return render(request, "icps/icps_average_report.html",
                      {'employees': users, 'title': title, 'start_date': start_date,
                       'end_date': end_date, 'summary_ready': ready})

    elif request.method == 'POST':
        summary_ready = True
        start_date = request.POST['date']
        end_date = request.POST["date2"]
        title = f"From {dt.datetime.strptime(start_date, '%Y-%m-%d').strftime('%d %b, %Y')} to " \
                f"{dt.datetime.strptime(end_date, '%Y-%m-%d').strftime('%d %b, %Y')}"

        access = ICPSAccess.objects.filter(date__range=(start_date, end_date)).values_list('date', 'name') \
            .distinct().order_by('-date', 'name')

        for i in access:
            employee_last_out = ICPSAccess.objects.filter(date=i[0], name=i[1]).latest('created_at')  # user time out
            employee_first_in = ICPSAccess.objects.filter(date=i[0], name=i[1]).earliest('created_at')  # user time in
            employee_last_out.timein = str(employee_first_in.created_at.time())
            employee_last_out.timeout = str(employee_last_out.created_at.time())

            # # user time out and time in difference
            t1 = dt.datetime.strptime(employee_last_out.timein, "%H:%M:%S")
            t2 = dt.datetime.strptime(employee_last_out.timeout, "%H:%M:%S")
            t = (t2 - t1)

            # checking and deducting breaktime
            if (t.total_seconds() / 3600) >= 1:
                t = (t2 - t1) - timedelta(hours=1)
            else:
                t = (t2 - t1)

            # create unique attendances in not created
            ICPSAccessUniqueAttendance.objects.get_or_create(date=str(employee_last_out.created_at.date()),
                                                             name=employee_last_out.name,
                                                             timein=str(employee_first_in.created_at.time()),
                                                             timeout=str(employee_last_out.created_at.time()),
                                                             w_hours=str(t)
                                                             )

        users = ICPSEmployee.objects.all().order_by('name')
        index = 0
        for i in users:
            index += 1
            i.comment = index  # Numbering in order

            # Finding average values
            average_time_in = ICPSAccessUniqueAttendance.objects.filter(date__range=(start_date, end_date),
                                                                        name=i.name).aggregate(
                avg_time_in=Avg('timein'))
            average_time_out = ICPSAccessUniqueAttendance.objects.filter(date__range=(start_date, end_date),
                                                                         name=i.name).aggregate(
                avg_time_out=Avg('timeout'))
            average_hours = ICPSAccessUniqueAttendance.objects.filter(date__range=(start_date, end_date),
                                                                      name=i.name).aggregate(avg_hours=Avg('w_hours'))

            i.av_timein = convert_to_hour_minute((average_time_in.get('avg_time_in', None)))
            i.av_timeout = convert_to_hour_minute((average_time_out.get('avg_time_out', None)))
            i.total = convert_to_hour_minute((average_hours.get('avg_hours', None)))

    return render(request, "icps/icps_average_report.html", {'employees': users, 'title': title,
                                                             'start_date': start_date, 'end_date': end_date,
                                                             'summary_ready': summary_ready})


# View detailed report of average attendance
@login_required(login_url='login')
def icps_employee_average_breakdown(request, start_date, end_date, title):
    employees = ICPSAccessUniqueAttendance.objects.filter(date__range=(start_date, end_date)).distinct().order_by(
        '-date')
    for i in employees:
        i.timein = convert_to_hour_minute(i.timein)
        i.timeout = convert_to_hour_minute(i.timeout)
        i.w_hours = convert_to_hour_minute(i.w_hours)
    return render(request, "icps/icps_avg_breakdown_report.html", {'employees': employees, 'title': title})


# User Attendance Breakdown
@login_required(login_url='login')
def user_attendance_breakdown(request, pk, title, start_date, end_date):
    employee = ICPSAccessUniqueAttendance.objects.filter(date__range=(start_date, end_date), name=pk). \
        distinct().order_by('-date')

    for i in employee:
        i.w_hours = convert_to_hour_minute(i.w_hours)

    json_data = {'employee': employee, 'title': title, 'user': pk}
    return render(request, 'icps/user_attendance_breakdown.html', json_data)

